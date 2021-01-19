import openpyxl

wb = openpyxl.Workbook() #Create new blank workbook object called wb

print(wb.sheetnames)   # prints names of the sheets in this workbook object

current_sheet_to_work_on = wb['Sheet'] # assigns the sheet to  a variable called 'current_sheet_to_work_on'
print(current_sheet_to_work_on)
wb.create_sheet()      #create a new sheet1
wb.create_sheet()      #create a new sheet2
print(wb.sheetnames)   #prints all the sheets in the workbook

current_sheet_to_work_on['A1'] = 89   # assign integer value 89 to cell 'A1'
wb.create_sheet(title= 'my first sheet', index= 0)
print(wb.sheetnames)
wb.save(filename='JudyTest2.xlsx')
path = r'C:\Users\Burudani\Documents\JudyTest4.xlsx' # use 'shft +rightclick', then select 'copy as path'
wb.save(filename=path)
