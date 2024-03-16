import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font
import json
import numpy as np
import tkinter as tk
from tkinter import filedialog
import datetime as dt
from datetime import datetime
from openpyxl.styles import PatternFill, Border, Side, Alignment
import os
from openpyxl.utils import get_column_letter


# Directory to check and save the Excel file
directory = r"G:/My Drive/Finances/"
base_filename = 'Example_Excel_Output_'
extension = '.xlsx'
current_date = datetime.now().strftime('%Y-%m-%d')

# Initialize version to 0 for each new day, this makes sure, any produced file is unique
version = 0
file_exists = True

# Create a tkinter root window
root = tk.Tk()

# Open a file dialog window to select an Excel (.xlsx) file
file_path = filedialog.askopenfilename(filetypes=[('Excel files', '*.xlsx')])

# Read the Excel file
df = pd.read_excel(file_path)


# Delete the unwanted 3rd and 4th columns (they have unecessary information in them)
df.drop(df.columns[[2, 3]], axis=1, inplace=True)


# Rename the remaining columns with new headers
df.columns = ['Date', 'Amount', 'Description']

# Ensure the 'DateTime' column is of datetime type
df['Date'] = pd.to_datetime(df['Date'])

# Sort the DataFrame based on the DateTime column from oldest to newest
df.sort_values(by='Date', inplace=True)

# Attempt to find the highest version for today
while file_exists:
    file_path2 = os.path.join(directory, f'{base_filename}{current_date}_v{version}{extension}')
    file_exists = os.path.exists(file_path2)
    if file_exists:
        # If file exists, increment version and check again
        version += 1
    else:
        # Save DataFrame to Excel file with the final file_path
        df.to_excel(file_path2, index=False)
        print(f"DataFrame saved to {file_path2}")
        break

excel_path = file_path2


# Load the workbook and sheet
wb = openpyxl.load_workbook(excel_path)
sheet = wb.active


# Define color fills, border style, and wrap text with top and left alignment
aqua_accent_5 = PatternFill(start_color="4AACC5", end_color="4AACC5", fill_type="solid")  # Aqua, Accent 5
lighter_60 = PatternFill(start_color="99D9EA", end_color="99D9EA", fill_type="solid")  # Lighter 60%
lighter_80 = PatternFill(start_color="E5F5FD", end_color="E5F5FD", fill_type="solid")  # Lighter 80%
top_left_alignment = Alignment(horizontal='left', vertical='top', wrapText=True)  # Updated alignment

# Define border style
thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

# Apply styles
for row in sheet.iter_rows(min_row=1, max_col=sheet.max_column, max_row=sheet.max_row):
    for cell in row:
        if cell.row == 1:
            cell.fill = aqua_accent_5
        elif cell.row == 2:
            cell.fill = lighter_60
        elif cell.row == 3:
            cell.fill = lighter_80
        else:
            # Alternate between lighter 60% and lighter 80% starting from the 4th row
            cell.fill = lighter_60 if (cell.row - 2) % 2 == 0 else lighter_80
        # Apply border and wrap text to every cell
        cell.border = thin_border
        cell.alignment = top_left_alignment  # Apply top and left alignment

# Set individual column widths
column_widths = [19, 10, 70, 40]
for i, col in enumerate(sheet.columns, start=1):
    sheet.column_dimensions[openpyxl.utils.get_column_letter(i)].width = column_widths[i-1]

# Enable filter for the column headers
sheet.auto_filter.ref = f"A1:{get_column_letter(sheet.max_column)}1"
sheet.freeze_panes = 'A2'  # Freeze the top row

# Save the workbook
wb.save(excel_path)
print ("Done Formatting Data")

#df.to_excel(r"Finances/fileTester_v0.xlsx", index = False)
