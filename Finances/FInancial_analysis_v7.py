import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font
import json
import tkinter as tk
from tkinter import filedialog
import datetime as dt
from datetime import datetime
from openpyxl.styles import PatternFill, Border, Side
import os

# Directory to check and save the Excel file
directory = r"G:/My Drive/Finances/"
base_filename = 'Latest_Financial_Expenditure_Statement_'
extension = '.xlsx'
current_date = datetime.now().strftime('%Y-%m-%d')

# Initialize version to 0 for each new day
version = 0
file_exists = True

# Load the lookup table from the JSON file
with open(r'C:\Users\Burudani\Documents\mainPythonFolder_v1\JSON\payDescriptionConverter_v2.json', 'r') as f:
    lookup_table = json.load(f)

# Create a tkinter root window
root = tk.Tk()

# Hide the root window
root.withdraw()

# Open a file dialog window to select the CSV file
file_path = filedialog.askopenfilename(filetypes=[('CSV files', '*.csv')])

# Load the CSV file into a pandas dataframe
df = pd.read_csv(file_path)
# Load the Excel file into a Pandas dataframe
#df = pd.read_excel(r"G:\My Drive\Finances\Latest_Statement_April_2023.xlsx")

# Delete the unwanted 3rd and 4th columns (they have unecessary information in them)
df.drop(df.columns[[2, 3]], axis=1, inplace=True)


# Rename the remaining columns with new headers
df.columns = ['date', 'amount', 'Description']

# Define the column to search and the column to create
search_col = 'Description'
create_col = 'Category'


# Function to find the category based on description
def find_category(description, lookup_table):
    description = description.lower()
    for category, keys in lookup_table.items():
        if category not in ['values']:  # Exclude the 'values' key
            for key in keys:
                if key.lower() in description:
                    # Return the corresponding value for the category
                    return lookup_table['values'][{
                        "credit_cards_&_Loans": "credit_cards_&_Loans",
                        "power_bill": "power_bill",
                        "clothes_and_hair": "clothes_and_hair",
                        "overdraft_fee": "overdraft_fee",
                        "fast_food": "fast_food",
                        "entertainment": "entertainment",
                        "education": "education",
                        "club_&_going_out": "club_&_going_out",
                        "fitness": "fitness",
                        "groceries": "groceries",
                        "car_care": "car_care",
                        "shopping": "shopping",
                        "salary": "salary",
                        "streaming_movies_and_music": "streaming_movies_and_music",
                        "alcohol": "alcohol",
                        "to_be_ignored": "to_be_ignored",
                        "travel": "travel",
                        "phone_bill": "phone_bill",
                        "rent_&_utils": "rent_&_utils",
                        "investment": "investment",
                        "investment_withdrawals": "investment_withdrawals",
                        "health": "health",
                        "rent_&_utils": "rent_&_utils"
                    }[category]]
    return "NO CATEGORY"

# Loop through each row and assign category based on description
for index, row in df.iterrows():
    category = find_category(row[search_col], lookup_table)
    df.loc[index, create_col] = category
# Get the current date in the format YYYY-MM-DD
#current_date = dt.datetime.now().strftime('%Y-%m-%d')

# Attempt to find the highest version for today
while file_exists:
    file_path2 = os.path.join(directory, f'{base_filename}{current_date}_v{version}{extension}')
    file_exists = os.path.exists(file_path2)
    if file_exists:
        # If file exists, increment version and check again
        version += 1
    else:
        # Save DataFrame to Excel file with the final file_path
        df.to_excel(file_path2, index=False)
        print(f"DataFrame saved to {file_path2}")
        break

excel_path = file_path2

# Load the workbook and sheet
# Load the workbook and sheet
wb = openpyxl.load_workbook(excel_path)
sheet = wb.active

# Define color fills
aqua_accent_5 = PatternFill(start_color="4AACC5", end_color="4AACC5", fill_type="solid")  # Aqua, Accent 5
lighter_60 = PatternFill(start_color="99D9EA", end_color="99D9EA", fill_type="solid")  # Lighter 60%
lighter_80 = PatternFill(start_color="E5F5FD", end_color="E5F5FD", fill_type="solid")  # Lighter 80%

# Define border style
thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

# Apply styles
for row in sheet.iter_rows(min_row=1, max_col=sheet.max_column, max_row=sheet.max_row):
    for cell in row:
        if cell.row == 1:
            cell.fill = aqua_accent_5
        elif cell.row == 2:
            cell.fill = lighter_60
        elif cell.row == 3:
            cell.fill = lighter_80
        else:
            # Alternate between lighter 60% and lighter 80% starting from the 4th row
            cell.fill = lighter_60 if (cell.row - 2) % 2 == 0 else lighter_80
        # Apply border to every cell
        cell.border = thin_border

# Save the workbook
wb.save(excel_path)
print ("Done with formatting")