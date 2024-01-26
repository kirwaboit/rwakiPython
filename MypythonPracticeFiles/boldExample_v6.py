import pandas as pd
from openpyxl.styles import PatternFill, Font,Border, Side

data = {'Name': ['Karan james', 'Rohit sing', 'Sahil nation', 'Aryan land'],
        'Age': [23, 22, 21, 24],
        'University': ['BHU', 'JNU', 'DU', 'BHU'],}

df = pd.DataFrame(data)

# Specify the file path where you want to save the Excel file
excel_file_path = r'C:\Users\Burudani\Documents\mainPythonFolder_v1\SpreadsheetFiles\boldExample_v83.xlsx'

# Use ExcelWriter to create the Excel file
with pd.ExcelWriter(excel_file_path, engine='openpyxl') as writer:
    df.to_excel(writer, index=False, sheet_name='Sheet1')

    # Access the workbook and the active sheet
    workbook = writer.book
    sheet = writer.sheets['Sheet1']

    # Remove existing filters (reset filters)
    sheet.auto_filter.ref = None

    # Apply filters to each column
    for col_num, col_letter in enumerate(sheet.iter_cols(min_col=1, max_col=len(df.columns)), start=1):
        sheet.auto_filter.ref = sheet.dimensions

    # Apply row styling
    for row_num, row in enumerate(sheet.iter_rows(min_row=1, max_row=1), start=1):
        for cell in row:
            fill_color = "5A9CCE"
            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        #     cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = Font(color="FFFFFF",name="Calibri")

    for row_num, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        if row_num % 2 == 0: # check even number rows
            fill_color = "BCD8EE"  # Blue, Accent 1, Lighter 60%   (LIGHT)
            for cell in row:
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                cell.font = Font(color="000000",name="Calibri")
        else:
            fill_color = "DDEBF6"  # Blue, Accent 1, Lighter 80%    (LIGHTER)
            for cell in row:
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                cell.font = Font(color="000000",name="Calibri")

    # Apply borders to the entire table  , white
    border = Border(left=Side(style='thin', color='FFFFFF'),
                    right=Side(style='thin', color='FFFFFF'),
                    top=Side(style='thin', color='FFFFFF'),
                    bottom=Side(style='thin', color='FFFFFF'))

    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.border = border   


    # Set column widths individually
    column_widths = [15, 10, 15]  # Specify the width for each column (adjust as needed)

    for i, width in enumerate(column_widths, start=1):
        sheet.column_dimensions[chr(64 + i)].width = width 
    
    # # Check the first column for the word "nation" and bold only that word in each cell
    # for row_num in range(2, sheet.max_row + 1):
    #     cell = sheet.cell(row=row_num, column=1)
    #     if "nation" in cell.value.lower():
    #         # Split the cell value by spaces and bold only the word "nation"
    #         words = cell.value.split()
    #         new_value = ""
    #         for word in words:
    #             if word.lower() == "nation":
    #                 # new_value += f'<b>{word}</b> '
    #                 new_value += f'<b>{word}</b> '
    #             else:
    #                 new_value += f"{word} "
    #         cell.value = new_value.strip()

print(f"DataFrame has been saved to {excel_file_path} with specified row styling.")
