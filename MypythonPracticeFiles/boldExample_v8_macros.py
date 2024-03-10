import pandas as pd
from openpyxl.styles import PatternFill, Font, Border, Side
import xlwings as xw

# Specify the path to the source workbook containing the VBA macro
source_wb_path = r'C:\Users\Burudani\Documents\mainPythonFolder_v1\SpreadsheetFiles\boldExample_v66_macroEdition.xlsm'

# Specify the path to the destination workbook where the DataFrame is saved with new data
excel_file_path = r'C:\Users\Burudani\Documents\mainPythonFolder_v1\SpreadsheetFiles\boldExample_v90.xlsx'

# Specify the name of the VBA macro
macro_name = "TypeName.RunPythonScript"

data = {'Name': ['Karan james', 'Rohit sing', 'Sahil nation', 'Aryan land'],
        'Age': [23, 22, 21, 24],
        'University': ['BHU', 'JNU', 'DU', 'BHU'],}

df = pd.DataFrame(data)

# Use ExcelWriter to create the Excel file with new data
with pd.ExcelWriter(excel_file_path, engine='openpyxl') as writer:
    df.to_excel(writer, index=False, sheet_name='Sheet1')

    # Access the workbook and the active sheet
    workbook = writer.book
    sheet = writer.sheets['Sheet1']

    # Your existing styling code...

    # Apply borders to the entire table, white
    border = Border(left=Side(style='thin', color='FFFFFF'),
                    right=Side(style='thin', color='FFFFFF'),
                    top=Side(style='thin', color='FFFFFF'),
                    bottom=Side(style='thin', color='FFFFFF'))

    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.border = border   

    # Set column widths individually
    column_widths = [15, 10, 15]  # Specify the width for each column (adjust as needed)

    for i, width in enumerate(column_widths, start=1):
        sheet.column_dimensions[chr(64 + i)].width = width

# Call the VBA script using xlwings
xl_app = xw.App(visible=False)
xl_wb = xl_app.books.open(source_wb_path)
xl_app.api.Run("TypeName.RunPythonScript")
xl_wb.save()
xl_wb.close()
xl_app.quit()

print(f"DataFrame has been saved to {excel_file_path} with specified row styling, and the VBA script has been executed from {source_wb_path}.")
